from flask import Blueprint, request, send_file, jsonify
from openpyxl import load_workbook
import os
import tempfile

bp = Blueprint('generate', __name__, url_prefix='/api/generate')

@bp.route('/excel', methods=['POST'])
def generate_excel():
    try:
        students = request.json.get("students", [])
        if not students:
            return jsonify({"error": "No student data received"}), 400

        # Load template
        template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'immersion_template.xlsx')
        wb = load_workbook(template_path)

        # Expected departments
        sheet_map = {
            "PRODUCTION": wb["PRODUCTION"],
            "SUPPORT": wb["SUPPORT"],
            "TECHNICAL": wb["TECHNICAL"]
        }

        # Get first student's values for direct cell assignment
        first_student = students[0] if students else {}
        immersion_date = first_student.get("date_of_immersion", "")
        batch = first_student.get("batch", "")
        school = first_student.get("school", "")

        # Put batch in M8, school in O8, date of immersion in M9 for all sheets
        for ws in wb.worksheets:
            ws['H8'] = batch
            ws['S8'] = school
            ws['S9'] = immersion_date

        # Start inserting from row 12 (row 11 = header)
        row_counter = {
            "PRODUCTION": 10,
            "SUPPORT": 10,
            "TECHNICAL": 10
        }

        for s in students:
            dept_raw = s.get("department", "").strip().upper()
            # Map department values
            if dept_raw == "IT":
                dept = "TECHNICAL"
            elif dept_raw == "PRODUCTION":
                dept = "PRODUCTION"
            else:
                dept = "SUPPORT"

            ws = sheet_map[dept]
            row = row_counter[dept]

            # Write data: 1G to 11G in columns H to R
            ws[f'B{row}'] = s.get("last_name", "")
            ws[f'C{row}'] = s.get("first_name", "")
            ws[f'D{row}'] = s.get("middle_name", "")
            ws[f'E{row}'] = s.get("strand", "")
            ws[f'F{row}'] = s.get("department", "")
            
            ws[f'H{row}'] = s.get("1G", "")
            ws[f'I{row}'] = s.get("2G", "")
            ws[f'J{row}'] = s.get("3G", "")
            ws[f'K{row}'] = s.get("4G", "")
            ws[f'L{row}'] = s.get("5G", "")
            ws[f'M{row}'] = s.get("6G", "")
            ws[f'N{row}'] = s.get("7G", "")
            ws[f'O{row}'] = s.get("8G", "")
            ws[f'P{row}'] = s.get("9G", "")
            ws[f'Q{row}'] = s.get("10G", "")
            ws[f'R{row}'] = s.get("11G", "")

            row_counter[dept] += 1

        # Save to temp file
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, "generated_immersion_report.xlsx")
        wb.save(output_path)

        return send_file(output_path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500